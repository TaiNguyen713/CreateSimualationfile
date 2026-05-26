"""
auto_create_config.py — Auto-generate Vehicle_infor-style config files (Excel)
by filtering and mapping data from the master NWS and DTC databases.

Workflow
--------
Step 1 – Minimal Vehicle Selection
  1a. Scan NWS Ymme filtered to CAN-based protocols (all InnovaGroups).
  1b. Run an online greedy set-cover that simultaneously selects YMMEs and
      assigns profiles:
        - In each iteration, simulate the effective profiles each remaining
          YMME would contribute (given already-committed globally_used state)
          and pick the YMME with the highest new coverage.
        - Commit: build system entries for that YMME, update globally_used.
      This ensures every CAN profile is actually present in at least one
      output file, with minimum duplicate assignments.
  1c. Profile assignment rules (applied during commit):
        - SubSystem rows (SubSystem != ""):  keep all — each subsystem has
          its own profile identity.
        - Non-subsystem rows (SubSystem == "") with multiple profiles:
          pick the first profile (NWS file order) not yet in globally_used;
          fall back to the first if all are already claimed.
  Each selected vehicle's config includes ALL its CAN systems (all
  InnovaGroups).  Vehicles that have fewer than MIN_SYSTEMS unique systems
  in the NWS database are accepted as-is — no artificial padding.

Step 2 – DTC Mapping (unchanged)
  For each system entry, join to the DTC database by MsgID/ECUID.
  Systems with no DTC match get a placeholder row with DTC = "0".

Usage
-----
    python auto_create_config.py              # writes to config/auto_configs/
    python auto_create_config.py out/dir      # custom output directory
"""

from __future__ import annotations

import argparse
import logging
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd
import polars as pl

# ---------------------------------------------------------------------------
# Paths and constants
# ---------------------------------------------------------------------------

NWS_DIR            = Path('data/NWS')
DTC_DIR            = Path('data/DTC')
DEFAULT_OUTPUT_DIR = Path('config/auto_configs')

DEFAULT_VIN    = '11111111111111111'
DEFAULT_STATUS = '00'



# Default protocol filter — passed to generate_all_configs().
# Pass None (or --all-protocols on the CLI) to include every protocol.
DEFAULT_PROTOCOLS: frozenset[str] | None = frozenset({'PROTOCOL_CAN_UDS', 'PROTOCOL_CAN'})

YMME_COLS   = ['Year', 'Manufacturer', 'Make', 'Model', 'Engine']
MIN_SYSTEMS = 10

_NAN = {'', 'nan', 'NaN'}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_xlsx(folder: Path) -> Path:
    files = sorted(folder.rglob('*.xlsx'))
    if not files:
        raise FileNotFoundError(f'No .xlsx files found in {folder}')
    if len(files) > 1:
        logger.warning('Multiple .xlsx files in %s -- using %s', folder, files[0].name)
    return files[0]


def _pd_to_pl(df: pd.DataFrame) -> pl.DataFrame:
    df = df.fillna('').astype(str)
    return pl.from_dict({col: df[col].tolist() for col in df.columns})


def _safe_name(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', text).strip()


def _ymme_tag(ymme: dict[str, str]) -> str:
    return _safe_name(
        f"{ymme['Year']}_{ymme['Manufacturer']}_{ymme['Make']}_"
        f"{ymme['Model']}_{ymme['Engine']}"
    )


def _ymme_key(row: dict) -> tuple[str, ...]:
    return tuple(row[c] for c in YMME_COLS)


def _key_to_ymme(key: tuple[str, ...]) -> dict[str, str]:
    return dict(zip(YMME_COLS, key))


# ---------------------------------------------------------------------------
# Load NWS Ymme (includes SubSystem)
# ---------------------------------------------------------------------------

def load_nws_ymme(nws_file: Path) -> pl.DataFrame:
    wanted = YMME_COLS + ['System', 'SubSystem', 'MsgID/ECUID', 'InnovaGroup']
    df = pd.read_excel(nws_file, sheet_name='Ymme', skiprows=[0, 2, 3], usecols=wanted)
    logger.info('Loaded NWS Ymme: %d rows', len(df))
    return _pd_to_pl(df)


# ---------------------------------------------------------------------------
# Load NWS Profile
# ---------------------------------------------------------------------------

def load_nws_profile(nws_file: Path) -> pl.DataFrame:
    df = pd.read_excel(
        nws_file, sheet_name='Profile', skiprows=[0, 2, 3],
        usecols=['MsgID/ECUID', 'Protocol'],
    )
    df = df.drop_duplicates(subset='MsgID/ECUID')
    logger.info('Loaded NWS Profile: %d unique profiles', len(df))
    return _pd_to_pl(df)


# ---------------------------------------------------------------------------
# Filtered views
# ---------------------------------------------------------------------------

def filter_nws(
    ymme: pl.DataFrame,
    profile: pl.DataFrame,
    protocols: frozenset[str] | None = DEFAULT_PROTOCOLS,
) -> pl.DataFrame:
    """
    Join *ymme* with *profile* on MsgID/ECUID.

    Parameters
    ----------
    protocols:
        Set of Protocol values to keep.  Pass ``None`` to include every
        protocol in the NWS Profile sheet (no filtering).
    """
    if protocols:
        prof_filtered = profile.filter(pl.col('Protocol').is_in(list(protocols)))
    else:
        prof_filtered = profile
    filtered = ymme.join(prof_filtered, on='MsgID/ECUID', how='inner')
    proto_desc = ', '.join(sorted(protocols)) if protocols else 'ALL'
    logger.info(
        'Protocol filter [%s]: %d rows | %d profiles | %d YMMEs',
        proto_desc,
        len(filtered),
        filtered['MsgID/ECUID'].n_unique(),
        filtered.select(YMME_COLS).n_unique(),
    )
    return filtered


# ---------------------------------------------------------------------------
# Core helpers: profile simulation and assignment
# ---------------------------------------------------------------------------

def _build_rows_by_ymme(filtered: pl.DataFrame) -> dict[tuple, list[dict]]:
    """Index filtered rows by YMME key, preserving NWS file order."""
    rows_by_ymme: dict[tuple, list[dict]] = defaultdict(list)
    for row in filtered.iter_rows(named=True):
        rows_by_ymme[_ymme_key(row)].append(row)
    return dict(rows_by_ymme)


def _effective_profiles(rows: list[dict], globally_used: set[str]) -> set[str]:
    """
    Simulate which profiles would be assigned to this YMME given *globally_used*,
    without modifying state.

    - SubSystem rows: contribute their profile unconditionally.
    - Non-subsystem rows: contribute the first profile not in globally_used,
      or the first profile if all are already used.
    """
    by_system: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_system[r['System']].append(r)

    profiles: set[str] = set()
    for sys_rows in by_system.values():
        sub_rows  = [r for r in sys_rows if r['SubSystem'] not in _NAN]
        main_rows = [r for r in sys_rows if r['SubSystem'] in _NAN]

        for r in sub_rows:
            profiles.add(r['MsgID/ECUID'])

        if main_rows:
            main_profiles = list(dict.fromkeys(r['MsgID/ECUID'] for r in main_rows))
            chosen = next((p for p in main_profiles if p not in globally_used), main_profiles[0])
            profiles.add(chosen)

    return profiles


def _commit_entries(rows: list[dict], globally_used: set[str]) -> list[dict[str, str]]:
    """
    Build system entries for one YMME, updating *globally_used* in-place.

    - SubSystem rows: include each unique profile (subsystem identity).
    - Non-subsystem rows: pick first profile not in globally_used; add to it.
    """
    by_system: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_system[r['System']].append(r)

    entries: list[dict[str, str]] = []

    for system, sys_rows in by_system.items():
        sub_rows  = [r for r in sys_rows if r['SubSystem'] not in _NAN]
        main_rows = [r for r in sys_rows if r['SubSystem'] in _NAN]

        seen_sub: set[str] = set()
        for r in sub_rows:
            p = r['MsgID/ECUID']
            if p not in seen_sub:
                entries.append({'System': system, 'MsgID/ECUID': p, 'Protocol': r.get('Protocol', '')})
                seen_sub.add(p)

        if main_rows:
            main_profiles = list(dict.fromkeys(r['MsgID/ECUID'] for r in main_rows))
            chosen = next((p for p in main_profiles if p not in globally_used), main_profiles[0])
            globally_used.add(chosen)
            chosen_row = next(r for r in main_rows if r['MsgID/ECUID'] == chosen)
            entries.append({'System': system, 'MsgID/ECUID': chosen, 'Protocol': chosen_row.get('Protocol', '')})

    return entries


# ---------------------------------------------------------------------------
# Online greedy set-cover with simultaneous profile assignment
# ---------------------------------------------------------------------------

def greedy_cover_and_assign(
    rows_by_ymme: dict[tuple, list[dict]],
    all_profiles: set[str],
) -> tuple[list[tuple], dict[tuple, list[dict[str, str]]]]:
    """
    Select the minimum set of YMMEs (greedy) while simultaneously assigning
    profiles to systems.

    In each iteration, the YMME whose *effective* new coverage (accounting for
    globally_used) is largest is selected.  Ties are broken by YMME key for
    determinism.

    Returns
    -------
    selected : ordered list of YMME keys
    entries  : {ymme_key: list_of_{System, MsgID/ECUID}_dicts}
    """
    uncovered: set[str] = set(all_profiles)
    remaining: set[tuple] = set(rows_by_ymme.keys())
    globally_used: set[str] = set()

    selected: list[tuple] = []
    entries_map: dict[tuple, list[dict[str, str]]] = {}

    while uncovered and remaining:
        best_key:  tuple | None = None
        best_gain: int          = -1

        for key in remaining:
            gain = len(_effective_profiles(rows_by_ymme[key], globally_used) & uncovered)
            if gain > best_gain or (gain == best_gain and (best_key is None or key < best_key)):
                best_gain = gain
                best_key  = key

        if best_gain == 0 or best_key is None:
            logger.warning('Cannot cover %d profiles: %s', len(uncovered), sorted(uncovered)[:5])
            break

        committed = _commit_entries(rows_by_ymme[best_key], globally_used)
        covered_now = {e['MsgID/ECUID'] for e in committed}
        uncovered -= covered_now

        selected.append(best_key)
        entries_map[best_key] = committed
        remaining.discard(best_key)

    logger.info('Set-cover done: %d YMMEs cover %d profiles', len(selected), len(all_profiles))

    # ── Patch: force-add any profiles still uncovered into an existing vehicle ─
    # This can happen when every vehicle that owns profile P already committed a
    # different variant for the same non-subsystem system.  We append P as an
    # extra entry to the first selected vehicle that has it in its rows.
    covered = {e['MsgID/ECUID'] for elist in entries_map.values() for e in elist}
    still_missing = all_profiles - covered
    if still_missing:
        logger.info('Patching %d residual profiles into existing vehicles', len(still_missing))
        # Build profile → row lookup over all rows
        profile_to_row: dict[str, dict] = {}
        for rows in rows_by_ymme.values():
            for r in rows:
                profile_to_row.setdefault(r['MsgID/ECUID'], r)

        for pid in sorted(still_missing):
            row = profile_to_row.get(pid)
            if row is None:
                logger.warning('Profile %s not found in any YMME — skipping', pid)
                continue
            # Find a selected vehicle that also has rows for this profile's system
            target_key: tuple | None = None
            for key in selected:
                for r in rows_by_ymme[key]:
                    if r['MsgID/ECUID'] == pid:
                        target_key = key
                        break
                if target_key:
                    break
            if target_key is None:
                # Not in any selected vehicle — use the first YMME in the DB
                target_key = _ymme_key(row)
                if target_key not in entries_map:
                    entries_map[target_key] = []
                    selected.append(target_key)
            entries_map[target_key].append({'System': row['System'], 'MsgID/ECUID': pid, 'Protocol': row.get('Protocol', '')})
            logger.info('  Patched %s -> %s (System=%s)', pid, target_key[2], row['System'])

    return selected, entries_map


# ---------------------------------------------------------------------------
# Step 2a – Load DTC sheet
# ---------------------------------------------------------------------------

def load_dtc(dtc_file: Path) -> pl.DataFrame:
    df = pd.read_excel(
        dtc_file, sheet_name='DTC', skiprows=[0, 2, 3],
        usecols=['Option 3', 'SAE DTC', 'Innova Group'],
    )
    logger.info('Loaded DTC sheet: %d rows', len(df))
    return _pd_to_pl(df)


def prepare_dtc(dtc_raw: pl.DataFrame) -> pl.DataFrame:
    return (
        dtc_raw
        .rename({'Option 3': 'MsgID/ECUID', 'SAE DTC': 'DTC'})
        .filter(
            (pl.col('MsgID/ECUID') != '') & (pl.col('MsgID/ECUID') != 'nan') &
            (pl.col('DTC')         != '') & (pl.col('DTC')          != 'nan')
        )
        .select(['MsgID/ECUID', 'DTC'])
    )


# ---------------------------------------------------------------------------
# Step 2b/c – Map systems to NWS rows (DTC)
# ---------------------------------------------------------------------------

def map_dtc_to_nws(
    systems: list[dict[str, str]],
    dtc_clean: pl.DataFrame,
) -> pl.DataFrame:
    """
    Build NWS config rows for one YMME.

    *systems* is a list of {System, MsgID/ECUID} dicts.
    Every system gets at least one row; no DTC match -> DTC = "0".
    """
    msgids     = list({e['MsgID/ECUID'] for e in systems})
    dtc_subset = dtc_clean.filter(pl.col('MsgID/ECUID').is_in(msgids))

    rows: list[dict[str, str]] = []
    for entry in systems:
        system   = entry['System']
        msgid    = entry['MsgID/ECUID']
        protocol = entry.get('Protocol', '')
        dtcs     = dtc_subset.filter(pl.col('MsgID/ECUID') == msgid)['DTC'].to_list()

        if not dtcs:
            rows.append({'System': system, 'Profile': msgid, 'Protocol': protocol, 'DTC': '0', 'Status': DEFAULT_STATUS})
        else:
            for code in dtcs:
                rows.append({'System': system, 'Profile': msgid, 'Protocol': protocol, 'DTC': code, 'Status': DEFAULT_STATUS})

    return pl.DataFrame(
        rows,
        schema={'System': pl.String, 'Profile': pl.String, 'Protocol': pl.String, 'DTC': pl.String, 'Status': pl.String},
    )


# ---------------------------------------------------------------------------
# Build VIN_YMME row
# ---------------------------------------------------------------------------

def build_vin_ymme(ymme: dict[str, str]) -> dict[str, str]:
    return {
        'VIN':          DEFAULT_VIN,
        'Year':         ymme['Year'],
        'Manufacturer': ymme['Manufacturer'],
        'Make':         ymme['Make'],
        'Model':        ymme['Model'],
        'Engine':       ymme['Engine'],
    }


# ---------------------------------------------------------------------------
# Write one config file
# ---------------------------------------------------------------------------

def write_config(
    vin_ymme: dict[str, str],
    nws_rows: pl.DataFrame,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_vin = wb.create_sheet('VIN_YMME')
    vin_header = ['VIN', 'Year', 'Manufacturer', 'Make', 'Model', 'Engine']
    ws_vin.append(vin_header)
    ws_vin.append([vin_ymme[col] for col in vin_header])

    ws_nws = wb.create_sheet('NWS')
    nws_header = ['System', 'Profile', 'Protocol', 'DTC', 'Status']
    ws_nws.append(nws_header)
    for row in nws_rows.iter_rows(named=True):
        ws_nws.append([row[col] for col in nws_header])

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_all_configs(
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    protocols: frozenset[str] | None = DEFAULT_PROTOCOLS,
) -> None:
    """
    Generate one config Excel file per selected YMME.

    The vehicle list is the minimal greedy set-cover of all profiles in the
    NWS database (all InnovaGroups) that match *protocols*.  Pass
    ``protocols=None`` to include every protocol without filtering.

    Parameters
    ----------
    output_dir:
        Directory where config files are written.
    protocols:
        Set of Protocol values to include.  ``None`` = no filter (all protocols).
        Default: ``DEFAULT_PROTOCOLS`` (CAN-based only).
    """
    nws_file = _find_xlsx(NWS_DIR)
    dtc_file = _find_xlsx(DTC_DIR)
    logger.info('NWS database : %s', nws_file)
    logger.info('DTC database : %s', dtc_file)

    # ── Load source data ──────────────────────────────────────────────────────
    logger.info('=== Loading source databases ===')
    ymme_df    = load_nws_ymme(nws_file)
    profile_df = load_nws_profile(nws_file)
    dtc_raw    = load_dtc(dtc_file)

    nws_data  = filter_nws(ymme_df, profile_df, protocols)
    dtc_clean = prepare_dtc(dtc_raw)

    # ── Step 1: Online greedy set-cover + profile assignment ──────────────────
    logger.info('=== Step 1: Online greedy set-cover ===')
    rows_by_ymme = _build_rows_by_ymme(nws_data)
    all_profiles = {r['MsgID/ECUID'] for rows in rows_by_ymme.values() for r in rows}

    selected_keys, entries_map = greedy_cover_and_assign(rows_by_ymme, all_profiles)

    total = len(selected_keys)
    logger.info('Selected %d vehicles (from %d unique YMMEs, %d profiles to cover)',
                total, len(rows_by_ymme), len(all_profiles))

    output_dir.mkdir(parents=True, exist_ok=True)

    # ── Step 2: DTC mapping + write ───────────────────────────────────────────
    logger.info('=== Step 2: DTC mapping + writing ===')
    ok = skipped = 0

    for i, key in enumerate(selected_keys, 1):
        ymme = _key_to_ymme(key)
        try:
            entries  = entries_map[key]
            n_sys    = len({e['System'] for e in entries})

            vin_ymme = build_vin_ymme(ymme)
            nws_rows = map_dtc_to_nws(entries, dtc_clean)

            tag      = _ymme_tag(ymme)
            out_path = output_dir / f'{tag}.xlsx'
            write_config(vin_ymme, nws_rows, out_path)
            ok += 1

            if i <= 5 or i == total:
                logger.info('[%d/%d] %s %s %s -- %d systems',
                            i, total, ymme['Year'], ymme['Make'], ymme['Model'], n_sys)

        except Exception as exc:
            logger.error('FAILED [%d/%d] %s %s %s %s: %s',
                         i, total, ymme.get('Year'), ymme.get('Make'),
                         ymme.get('Model'), ymme.get('Engine'), exc)
            skipped += 1

        if i % 50 == 0:
            logger.info('Progress: %d/%d  (ok=%d  failed=%d)', i, total, ok, skipped)

    logger.info('Done. %d config files written to %s  (skipped=%d)', ok, output_dir, skipped)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Generate minimal vehicle config files covering all NWS profiles.',
    )
    parser.add_argument(
        'output_dir', nargs='?', type=Path, default=DEFAULT_OUTPUT_DIR,
        help=f'Output directory (default: {DEFAULT_OUTPUT_DIR})',
    )
    proto_group = parser.add_mutually_exclusive_group()
    proto_group.add_argument(
        '--protocols', metavar='P1,P2', type=lambda s: frozenset(s.split(',')),
        default=DEFAULT_PROTOCOLS,
        help=(
            'Comma-separated list of Protocol values to include '
            f'(default: {",".join(sorted(DEFAULT_PROTOCOLS))})'
        ),
    )
    proto_group.add_argument(
        '--all-protocols', dest='all_protocols', action='store_true',
        help='Include all protocols — no protocol filter applied',
    )
    args = parser.parse_args()

    protocols = None if args.all_protocols else args.protocols
    generate_all_configs(args.output_dir, protocols)
